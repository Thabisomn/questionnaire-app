import os
from flask import Flask, render_template, request, redirect, url_for, send_file, flash
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key-in-production")
app.jinja_env.globals.update(enumerate=enumerate)

DATABASE_URL = os.environ.get("DATABASE_URL")
TABLE_NAME = "questionnaire_stakeholder_responses"  # prefixed/specific so it never clashes with DLEMS tables

USE_POSTGRES = bool(DATABASE_URL)

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
    # Render's DATABASE_URL sometimes starts with postgres:// ; psycopg2 wants postgresql://
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
else:
    import sqlite3
    DB_FILE = "responses.db"


# ── Database helpers ────────────────────────────────────────────────────────
def get_db():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        return conn
    else:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        return conn


def run_query(query_pg, query_sqlite, params=(), fetch=None):
    """
    fetch: None (just execute/commit), 'one', or 'all'
    query_pg uses %s placeholders, query_sqlite uses ? placeholders
    """
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(query_pg if USE_POSTGRES else query_sqlite, params)
        result = None
        if fetch == "one":
            result = cur.fetchone()
        elif fetch == "all":
            result = cur.fetchall()
        conn.commit()
        return result
    finally:
        conn.close()


def init_db():
    if USE_POSTGRES:
        ddl = f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            id SERIAL PRIMARY KEY,
            submitted_at TEXT,
            consent TEXT,
            gender TEXT,
            age_group TEXT,
            district TEXT,
            stakeholder_category TEXT,
            education TEXT,
            b1_planning INTEGER, b2_decision_making INTEGER, b3_implementation INTEGER, b4_consulted INTEGER,
            b5_info_communicated INTEGER, b6_regular_updates INTEGER, b7_effective_communication INTEGER,
            b8_informed_decisions INTEGER, b9_feedback_considered INTEGER,
            b10_work_together INTEGER, b11_cooperation INTEGER, b12_share_responsibilities INTEGER, b13_partnerships INTEGER,
            b14_influence INTEGER, b15_leadership_roles INTEGER, b16_trained INTEGER, b17_capacity INTEGER,
            c1_continues_after_funding INTEGER, c2_maintains_infrastructure INTEGER, c3_meets_needs INTEGER, c4_systems_in_place INTEGER,
            d1_feel_included INTEGER, d2_trust_leadership INTEGER, d3_satisfied_participation INTEGER, d4_engagement_fair INTEGER,
            e1_lack_resources INTEGER, e2_power_imbalances INTEGER, e3_lack_information INTEGER, e4_weak_institutions INTEGER,
            f1_strong_leadership INTEGER, f2_training INTEGER, f3_good_communication INTEGER, f4_government_support INTEGER
        )"""
    else:
        ddl = f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            submitted_at TEXT, consent TEXT, gender TEXT, age_group TEXT, district TEXT,
            stakeholder_category TEXT, education TEXT,
            b1_planning INTEGER, b2_decision_making INTEGER, b3_implementation INTEGER, b4_consulted INTEGER,
            b5_info_communicated INTEGER, b6_regular_updates INTEGER, b7_effective_communication INTEGER,
            b8_informed_decisions INTEGER, b9_feedback_considered INTEGER,
            b10_work_together INTEGER, b11_cooperation INTEGER, b12_share_responsibilities INTEGER, b13_partnerships INTEGER,
            b14_influence INTEGER, b15_leadership_roles INTEGER, b16_trained INTEGER, b17_capacity INTEGER,
            c1_continues_after_funding INTEGER, c2_maintains_infrastructure INTEGER, c3_meets_needs INTEGER, c4_systems_in_place INTEGER,
            d1_feel_included INTEGER, d2_trust_leadership INTEGER, d3_satisfied_participation INTEGER, d4_engagement_fair INTEGER,
            e1_lack_resources INTEGER, e2_power_imbalances INTEGER, e3_lack_information INTEGER, e4_weak_institutions INTEGER,
            f1_strong_leadership INTEGER, f2_training INTEGER, f3_good_communication INTEGER, f4_government_support INTEGER
        )"""

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(ddl)
        conn.commit()
    finally:
        conn.close()


init_db()

PH = "%s" if USE_POSTGRES else "?"  # placeholder style


def ph_list(n):
    return ",".join([PH] * n)


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/", methods=["GET"])
def index():
    return render_template("form.html")


@app.route("/submit", methods=["POST"])
def submit():
    f = request.form

    consent = f.get("consent", "No")
    if consent != "Yes":
        flash("You must consent to participate in this study.", "warning")
        return redirect(url_for("index"))

    columns = [
        "submitted_at", "consent", "gender", "age_group", "district",
        "stakeholder_category", "education",
        "b1_planning", "b2_decision_making", "b3_implementation", "b4_consulted",
        "b5_info_communicated", "b6_regular_updates", "b7_effective_communication",
        "b8_informed_decisions", "b9_feedback_considered",
        "b10_work_together", "b11_cooperation", "b12_share_responsibilities", "b13_partnerships",
        "b14_influence", "b15_leadership_roles", "b16_trained", "b17_capacity",
        "c1_continues_after_funding", "c2_maintains_infrastructure", "c3_meets_needs", "c4_systems_in_place",
        "d1_feel_included", "d2_trust_leadership", "d3_satisfied_participation", "d4_engagement_fair",
        "e1_lack_resources", "e2_power_imbalances", "e3_lack_information", "e4_weak_institutions",
        "f1_strong_leadership", "f2_training", "f3_good_communication", "f4_government_support",
    ]

    data = (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        consent,
        f.get("gender"), f.get("age_group"), f.get("district"),
        f.get("stakeholder_category"), f.get("education"),
        f.get("b1"), f.get("b2"), f.get("b3"), f.get("b4"),
        f.get("b5"), f.get("b6"), f.get("b7"), f.get("b8"), f.get("b9"),
        f.get("b10"), f.get("b11"), f.get("b12"), f.get("b13"),
        f.get("b14"), f.get("b15"), f.get("b16"), f.get("b17"),
        f.get("c1"), f.get("c2"), f.get("c3"), f.get("c4"),
        f.get("d1"), f.get("d2"), f.get("d3"), f.get("d4"),
        f.get("e1"), f.get("e2"), f.get("e3"), f.get("e4"),
        f.get("f1"), f.get("f2"), f.get("f3"), f.get("f4"),
    )

    query = f"INSERT INTO {TABLE_NAME} ({', '.join(columns)}) VALUES ({ph_list(len(columns))})"
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(query, data)
        conn.commit()
    finally:
        conn.close()

    return render_template("thankyou.html")


@app.route("/admin")
def admin():
    conn = get_db()
    try:
        cur = conn.cursor()

        cur.execute(f"SELECT COUNT(*) FROM {TABLE_NAME}")
        total = list(cur.fetchone().values())[0] if USE_POSTGRES else cur.fetchone()[0]

        if USE_POSTGRES:
            cur.execute(f"SELECT COUNT(*) FROM {TABLE_NAME} WHERE submitted_at::date = CURRENT_DATE")
            today = list(cur.fetchone().values())[0]
            cur.execute(f"SELECT COUNT(*) FROM {TABLE_NAME} WHERE submitted_at::timestamp >= NOW() - INTERVAL '7 days'")
            this_week = list(cur.fetchone().values())[0]
        else:
            cur.execute(f"SELECT COUNT(*) FROM {TABLE_NAME} WHERE DATE(submitted_at) = DATE('now')")
            today = cur.fetchone()[0]
            cur.execute(f"SELECT COUNT(*) FROM {TABLE_NAME} WHERE submitted_at >= DATE('now', '-7 days')")
            this_week = cur.fetchone()[0]

        cur.execute(f"SELECT gender, COUNT(*) as cnt FROM {TABLE_NAME} GROUP BY gender")
        gender_rows = cur.fetchall()
        gender = {(r["gender"] if USE_POSTGRES else r["gender"]): (r["cnt"] if USE_POSTGRES else r["cnt"]) for r in gender_rows}

        cur.execute(f"SELECT district, COUNT(*) as cnt FROM {TABLE_NAME} GROUP BY district ORDER BY cnt DESC")
        districts = [(r["district"], r["cnt"]) for r in cur.fetchall()]

        cur.execute(f"SELECT stakeholder_category, COUNT(*) as cnt FROM {TABLE_NAME} GROUP BY stakeholder_category ORDER BY cnt DESC")
        categories = [(r["stakeholder_category"], r["cnt"]) for r in cur.fetchall()]

        cur.execute(f"SELECT submitted_at, gender, age_group, district, stakeholder_category FROM {TABLE_NAME} ORDER BY submitted_at DESC LIMIT 10")
        recent = cur.fetchall()
    finally:
        conn.close()

    now = datetime.now()
    hour = now.hour
    greeting = "Good morning" if hour < 12 else "Good afternoon" if hour < 17 else "Good evening"

    return render_template(
        "admin.html",
        total=total, today=today, this_week=this_week,
        gender=gender, districts=districts, categories=categories,
        recent=recent, greeting=greeting,
        now=now.strftime("%A, %d %B %Y"),
    )


@app.route("/download")
def download():
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM {TABLE_NAME} ORDER BY submitted_at DESC")
        rows = cur.fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    section_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        ("ID", "Meta"), ("Submitted At", "Meta"), ("Consent", "Meta"),
        ("Gender", "Section A"), ("Age Group", "Section A"), ("District", "Section A"),
        ("Stakeholder Category", "Section A"), ("Education Level", "Section A"),
        ("Involved in project planning", "B: Participation"),
        ("Community participates in decisions", "B: Participation"),
        ("Stakeholders in implementation", "B: Participation"),
        ("Community consulted regularly", "B: Participation"),
        ("Info clearly communicated", "B: Communication"),
        ("Regular updates received", "B: Communication"),
        ("Effective communication", "B: Communication"),
        ("Informed about decisions", "B: Communication"),
        ("Feedback considered", "B: Communication"),
        ("Work together in activities", "B: Collaboration"),
        ("Cooperation with leaders", "B: Collaboration"),
        ("Share responsibilities", "B: Collaboration"),
        ("Govt-NGO-Community partnerships", "B: Collaboration"),
        ("Influence over decisions", "B: Empowerment"),
        ("Community given leadership roles", "B: Empowerment"),
        ("Stakeholders trained", "B: Empowerment"),
        ("Local capacity to sustain", "B: Empowerment"),
        ("Continues after funding ends", "C: Sustainability"),
        ("Maintains infrastructure", "C: Sustainability"),
        ("Meets long-term needs", "C: Sustainability"),
        ("Systems to sustain project", "C: Sustainability"),
        ("Feel included in decisions", "D: Perceptions"),
        ("Trust project leadership", "D: Perceptions"),
        ("Satisfied with participation", "D: Perceptions"),
        ("Engagement is fair", "D: Perceptions"),
        ("Lack of resources limits participation", "E: Barriers"),
        ("Power imbalances affect participation", "E: Barriers"),
        ("Lack of info limits engagement", "E: Barriers"),
        ("Weak institutions affect sustainability", "E: Barriers"),
        ("Strong leadership improves engagement", "F: Enablers"),
        ("Training improves participation", "F: Enablers"),
        ("Good communication enhances success", "F: Enablers"),
        ("Government support improves sustainability", "F: Enablers"),
    ]

    sections = {}
    for col_idx, (label, section) in enumerate(headers, start=1):
        if section not in sections:
            sections[section] = [col_idx, col_idx]
        else:
            sections[section][1] = col_idx

    ws.append([""] * len(headers))
    for section, (start, end) in sections.items():
        cell = ws.cell(row=1, column=start, value=section)
        cell.fill = section_fill
        cell.font = header_font
        cell.alignment = center
        if start != end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    for col_idx, (label, _) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in rows:
        if USE_POSTGRES:
            ws.append(list(row.values()))
        else:
            ws.append(list(row))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"questionnaire_responses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
